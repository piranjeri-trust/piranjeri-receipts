"""
generate_report.py
Piranjeri Temples Family Trust — Collections Report Generator

Produces an Excel workbook with two sheets:
  1. Collections        — active (valid) receipts with totals by payment method
  2. Cancelled Receipts — voided receipts with cancellation details

Usage (called from app.py):
    generate_collections_report(
        month_data=month_active,
        month_label="April 2025",
        output_path=Path("reports/Collections_April_2025.xlsx"),
        cancelled_data=month_cancelled,   # list of cancelled receipt dicts
    )
"""

from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Colour palette ─────────────────────────────────────────────────────────
NAVY        = "003087"
SAFFRON     = "C45C00"
LIGHT_BLUE  = "D6E4F7"
LIGHT_RED   = "FFE4E1"
CANCELLED_ROW = "FFDCDC"
WHITE       = "FFFFFF"
DARK_GREY   = "444444"
HEADER_FONT_COLOR = "FFFFFF"

# ── Thin border helper ─────────────────────────────────────────────────────
THIN = Side(style="thin", color="CCCCCC")
def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fmt_date(iso: str) -> str:
    """Convert YYYY-MM-DD to DD MMM YYYY."""
    try:
        return datetime.strptime(iso, "%Y-%m-%d").strftime("%d %b %Y")
    except Exception:
        return iso


def _payment_label(method: str) -> str:
    return {
        "cash":          "Cash",
        "cheque":        "Cheque",
        "bank_transfer": "Bank Transfer",
    }.get(str(method).lower(), method)


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_row(ws, title: str, ncols: int, fill_hex: str = NAVY):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(bold=True, size=13, color=HEADER_FONT_COLOR)
    cell.fill      = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24


def _header_row(ws, row: int, headers: list, fill_hex: str = SAFFRON):
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font      = Font(bold=True, size=10, color=HEADER_FONT_COLOR)
        c.fill      = PatternFill("solid", fgColor=fill_hex)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
    ws.row_dimensions[row].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ACTIVE COLLECTIONS
# ══════════════════════════════════════════════════════════════════════════════
def _write_collections_sheet(ws, month_data: List[Dict], month_label: str):
    headers = [
        "S.No", "Receipt No.", "Donor Name", "Amount (Rs.)",
        "Payment Method", "Cheque No.", "Purpose",
        "Date of Credit", "Date of Issue", "Issued By"
    ]
    ncols = len(headers)
    col_widths = [6, 14, 28, 14, 16, 12, 28, 16, 16, 16]

    _title_row(ws, f"Piranjeri Temples Family Trust — Collections: {month_label}", ncols)
    _header_row(ws, 2, headers)
    _set_col_widths(ws, col_widths)

    # Payment-method subtotals
    method_totals: Dict[str, float] = {}
    grand_total = 0.0

    for idx, h in enumerate(month_data, start=1):
        row = idx + 2
        amt = float(h.get("amount", 0))
        pm  = _payment_label(h.get("payment", ""))
        method_totals[pm] = method_totals.get(pm, 0) + amt
        grand_total       += amt

        fill = PatternFill("solid", fgColor=LIGHT_BLUE if idx % 2 == 0 else WHITE)
        values = [
            idx,
            h.get("serial", ""),
            h.get("name", ""),
            amt,
            pm,
            h.get("cheque_number", ""),
            h.get("purpose", ""),
            _fmt_date(h.get("credit_date", "")),
            _fmt_date(h.get("issue_date", "")),
            h.get("user", ""),
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.border    = thin_border()
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if col == 4 else "center" if col in (1,5,6,8,9,10) else "left")
            if col == 4:
                c.number_format = '#,##0.00'

    # ── Totals section ─────────────────────────────────────────────
    totals_start = len(month_data) + 4

    ws.cell(row=totals_start - 1, column=1,
            value="Payment Method Totals").font = Font(bold=True, size=10, color=NAVY)

    sub_headers = ["Payment Method", "No. of Receipts", "Amount (Rs.)"]
    for col, label in enumerate(sub_headers, start=1):
        c = ws.cell(row=totals_start, column=col, value=label)
        c.font      = Font(bold=True, color=HEADER_FONT_COLOR)
        c.fill      = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border    = thin_border()

    method_count = {pm: sum(1 for h in month_data if _payment_label(h.get("payment","")) == pm)
                    for pm in method_totals}

    for r_off, (pm, total) in enumerate(sorted(method_totals.items()), start=1):
        r = totals_start + r_off
        for col, val in enumerate([pm, method_count[pm], total], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border    = thin_border()
            c.alignment = Alignment(horizontal="right" if col == 3 else "center")
            if col == 3:
                c.number_format = '#,##0.00'

    grand_row = totals_start + len(method_totals) + 1
    for col, val in enumerate(["GRAND TOTAL", len(month_data), grand_total], start=1):
        c = ws.cell(row=grand_row, column=col, value=val)
        c.font      = Font(bold=True, size=11, color=HEADER_FONT_COLOR)
        c.fill      = PatternFill("solid", fgColor=SAFFRON)
        c.border    = thin_border()
        c.alignment = Alignment(horizontal="right" if col == 3 else "center")
        if col == 3:
            c.number_format = '#,##0.00'


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — CANCELLED RECEIPTS
# ══════════════════════════════════════════════════════════════════════════════
def _write_cancelled_sheet(ws, cancelled_data: List[Dict], month_label: str):
    headers = [
        "S.No", "Receipt No.", "Donor Name", "Amount (Rs.)",
        "Payment Method", "Purpose", "Date of Issue",
        "Cancelled By", "Cancelled On", "Reason for Cancellation"
    ]
    ncols = len(headers)
    col_widths = [6, 14, 28, 14, 16, 24, 16, 16, 20, 36]

    _title_row(ws,
               f"Piranjeri Temples Family Trust — Cancelled Receipts: {month_label}",
               ncols, fill_hex="8B0000")
    _header_row(ws, 2, headers, fill_hex="C0392B")
    _set_col_widths(ws, col_widths)

    if not cancelled_data:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
        c = ws.cell(row=3, column=1, value="No cancelled receipts for this month.")
        c.font      = Font(italic=True, color=DARK_GREY)
        c.alignment = Alignment(horizontal="center")
        return

    cancelled_total = 0.0
    for idx, h in enumerate(cancelled_data, start=1):
        row  = idx + 2
        amt  = float(h.get("amount", 0))
        cancelled_total += amt
        fill = PatternFill("solid", fgColor=CANCELLED_ROW if idx % 2 == 1 else LIGHT_RED)

        # Format cancelled_at timestamp
        raw_ts = h.get("cancelled_at", "")
        try:
            cancelled_on = datetime.fromisoformat(raw_ts).strftime("%d %b %Y %H:%M")
        except Exception:
            cancelled_on = raw_ts

        values = [
            idx,
            h.get("serial", ""),
            h.get("name", ""),
            amt,
            _payment_label(h.get("payment", "")),
            h.get("purpose", ""),
            _fmt_date(h.get("issue_date", "")),
            h.get("cancelled_by", ""),
            cancelled_on,
            h.get("cancel_reason", ""),
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.border    = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 10),
                                    horizontal="right" if col == 4 else "center" if col in (1,5,7,8,9) else "left")
            if col == 4:
                c.number_format = '#,##0.00'

    # ── Cancelled total row ────────────────────────────────────────
    total_row = len(cancelled_data) + 4
    ws.cell(row=total_row, column=1,
            value=f"Total cancelled: {len(cancelled_data)} receipt(s)").font = Font(bold=True, color="8B0000")
    c = ws.cell(row=total_row, column=4, value=cancelled_total)
    c.font           = Font(bold=True, color="8B0000")
    c.number_format  = '#,##0.00'
    c.alignment      = Alignment(horizontal="right")
    ws.cell(row=total_row, column=3,
            value="Total amount voided →").font = Font(bold=True, color="8B0000")


# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def generate_collections_report(
    month_data: List[Dict],
    month_label: str,
    output_path: Path,
    cancelled_data: Optional[List[Dict]] = None,
):
    """
    Generate an Excel report with:
      - Sheet 1 "Collections"        : active receipts + payment-method totals
      - Sheet 2 "Cancelled Receipts" : voided receipts with cancellation details

    Parameters
    ----------
    month_data      : list of active receipt dicts for the selected month
    month_label     : human-readable month string, e.g. "April 2025"
    output_path     : Path where the .xlsx file will be saved
    cancelled_data  : list of cancelled receipt dicts (optional; defaults to [])
    """
    if cancelled_data is None:
        cancelled_data = []

    wb = openpyxl.Workbook()

    # Sheet 1 — Collections
    ws1 = wb.active
    ws1.title = "Collections"
    _write_collections_sheet(ws1, month_data, month_label)

    # Sheet 2 — Cancelled Receipts
    ws2 = wb.create_sheet("Cancelled Receipts")
    _write_cancelled_sheet(ws2, cancelled_data, month_label)

    # Metadata
    wb.properties.title   = f"Collections Report — {month_label}"
    wb.properties.creator = "Piranjeri Temples Family Trust"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
